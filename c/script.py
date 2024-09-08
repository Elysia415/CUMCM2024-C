import json
import os
import heapq
import copy
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import numpy as np
def greedysolve(season, muchan, pqall, descenddict, get_row_number, alldict, newdict):
    _ground = copy.deepcopy(ground)
    totalprofit = 0
    checkwhat = defaultdict(list)
    for place, plantdict in checkhvwhat.items():
        for plant in plantdict:
            checkwhat[place].append(int(plant))
    print(checkwhat)
    newhvwhat = defaultdict(lambda: defaultdict(int))
    while len(pqall) > 0:
        profit, plid, grid = pqall[0]
        profit = -profit
        planted = False
        if plid > 50:
            plid1, plid2 = newdict[plid]
            for i in range(len(_ground[grid])):
                pl1status = alldict[(plid1, grid)]
                pro1 = pl1status[0] * pl1status[1] * (1 + bonus[leveldict[plid1]]) - pl1status[2]
                depro1 = 0.5 * pl1status[0] * pl1status[1] * (1 + bonus[leveldict[plid1]]) - pl1status[2]
                pl2status = alldict[(plid2, grid)]
                pro2 = pl2status[0] * pl2status[1] * (1 + 0.1) - pl2status[2]
                depro2 = 0.5 * pl2status[0] * pl2status[1] * (1 + 0.1) - pl2status[2]

                cond1 = maxsold[plid1] / muchan[(plid1, grid)] <= limitminarea[grid]
                cond2 = maxsold[plid2] / muchan[(plid2, grid)] <= limitminarea[grid]
                if cond1 and cond2:  # 两个都到限额了
                    planted = True
                    heapq.heappop(pqall)
                    # 都到了第二个限额
                    if changed[(plid1, grid)] and changed[(plid2, grid)]:
                        break
                    # 至少一个是到第一个限额
                    else:
                        # 1到第一个限额
                        if not changed[(plid1, grid)]:
                            changed[(plid1, grid)] = 1
                            maxsold[plid1] += maxsold_2[plid1]
                        # 2到第一个限额
                        if not changed[(plid2, grid)]:
                            changed[(plid2, grid)] = 1
                            maxsold[plid2] += maxsold_2[plid2]
                        if not changed[(plid1, grid)] and not changed[(plid2, grid)]:
                            heapq.heappush(pqall, (-(depro1 + depro2) / 2, plid, grid))
                    break
                elif cond1:  # 只有1到了限额
                    if not changed[(plid1, grid)]:  # 是第一个限额
                        changed[(plid1, grid)] = 1
                        maxsold[plid1] += maxsold_2[plid1]
                        heapq.heappush(pqall, (-(depro1 + pro2) / 2, plid, grid))
                    break
                elif cond2:  # 只有2到了限额
                    if not changed[(plid2, grid)]:
                        changed[(plid2, grid)] = 1
                        maxsold[plid2] += maxsold_2[plid2]
                        heapq.heappush(pqall, (-(depro2 + pro1) / 2, plid, grid))
                    break
                place, area = _ground[grid][i]
                if plid1 in checkwhat[place] or plid2 in checkwhat[place]:
                    continue
                if area <= limitminarea[grid]:
                    continue
                planted = True
                plantarea = min(area, maxsold[plid1] / muchan[(plid1, grid)], maxsold[plid2] / muchan[(plid2, grid)])
                totalprofit += plantarea * profit
                _ground[grid][i][1] -= plantarea
                maxsold[plid1] -= plantarea * muchan[(plid1, grid)]
                maxsold[plid2] -= plantarea * muchan[(plid2, grid)]
                newhvwhat[place][plid1] += plantarea / 2
                newhvwhat[place][plid2] += plantarea / 2
        else:
            for i in range(len(_ground[grid])):
                if maxsold[plid] / muchan[(plid, grid)] <= limitminarea[grid]:
                    planted = True
                    heapq.heappop(pqall)
                    if not changed[(plid, grid)]:
                        changed[(plid, grid)] = 1
                        heapq.heappush(pqall, (-descenddict[(plid, grid)], plid, grid))
                        maxsold[plid] = maxsold_2[plid]
                    break
                place, area = _ground[grid][i]
                # 关于豆类的约束
                if hvpea[place] >= 2 and plid not in pealist:
                    continue
                # 关于上一季种植作物的约束
                if plid in checkwhat[place]:
                    continue
                if area <= limitminarea[grid]:
                    continue
                planted = True
                # 贪心,尽量种最多
                plantarea = min(area, maxsold[plid] / muchan[(plid, grid)])
                totalprofit += plantarea * profit
                _ground[grid][i][1] -= plantarea
                maxsold[plid] -= plantarea * muchan[(plid, grid)]
                newhvwhat[place][plid] += plantarea
        if not planted:
            heapq.heappop(pqall)
            continue
    # 更新豆类约束
    if season == 1:
        for place, plantdict in newhvwhat.items():
            p = 0
            for plid, plantarea in plantdict.items():
                if plid in pealist:
                    p = 1
                    break
            season1pea[place] = p
    else:
        for place, plantdict in newhvwhat.items():
            p = 0
            for plid, plantarea in plantdict.items():
                if plid in pealist:
                    p = 1
                    break
            season2pea[place] = p
        for place in hvpea.keys():
            if season1pea[place] or season2pea[place]:
                hvpea[place] = 0
            else:
                hvpea[place] += 1
    for place in newhvwhat:
        checkhvwhat[place] = newhvwhat[place]
    if season == 1:
        for grid in [1, 2, 3, 4, 5, 6]:
            for place, area in _ground[grid]:
                if place not in newhvwhat:
                    checkhvwhat[place] = defaultdict(int)
    sheetname = str(year + 1)
    sheet = workbook[sheetname]
    for place, plantdict in newhvwhat.items():
        for plid, plantarea in plantdict.items():
            col = get_column_letter(plid + 2)
            row = get_row_number[place]
            cell_address = f"{col}{row}"
            sheet[cell_address] = plantarea
    return totalprofit


workbook = load_workbook('result.xlsx')
with open("get_row_number1.json", 'r', encoding='utf-8') as f:
    raw_data = json.load(f)
get_row_number1 = {}
for row_dict in raw_data:
    number, place = row_dict.values()
    get_row_number1[place] = number
with open("get_row_number2.json", 'r', encoding='utf-8') as f:
    raw_data = json.load(f)
get_row_number2 = {}
for row_dict in raw_data:
    number, place = row_dict.values()
    get_row_number2[place] = number

with open("elasticity.json", 'r', encoding='utf-8') as f:
    raw_data = json.load(f)
elasticity = defaultdict(int)
for adict in raw_data:
    values = list(adict.values())
    elasticity[values[0]] = values[3]

vegetables = [17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37]
mushrooms = [38, 39, 40]
year = 2023
totalprofit = 0
directory = os.getcwd()
limitminarea = {1: 1, 2: 1, 3: 1, 4: 1, 5: 0.1, 6: 0.1}
while year < 2030:
    with open(os.path.join(directory, str(year), "1_1.json"), 'r', encoding='utf-8') as f:
        data1_1 = json.load(f)
    with open(os.path.join(directory, str(year), "1_2.json"), 'r', encoding='utf-8') as f:
        data1_2 = json.load(f)
    with open(os.path.join(directory, str(year), "2.json"), 'r', encoding='utf-8') as f:
        data2 = json.load(f)
    with open(os.path.join(directory, str(year), "3.json"), 'r', encoding='utf-8') as f:
        data3 = json.load(f)
    with open(os.path.join(directory, str(year), "4.json"), 'r', encoding='utf-8') as f:
        data4 = json.load(f)
    if year == 2023:
        with open(os.path.join(directory, str(year), "5.json"), 'r', encoding='utf-8') as f:
            data5 = json.load(f)
    if year != 2023:
        with open(os.path.join(directory, str(year), "checkhvwhat.json"), 'r', encoding='utf-8') as f:
            checkhvwhat = json.load(f)
    else:
        checkhvwhat = defaultdict(lambda: defaultdict(int))
        for data in data5:
            values = list(data.values())
            checkhvwhat[values[0]][values[1]] = 0

    # data1_1/1_2: 每一季（作物编号，地块类型编号）所对应的单价、亩产量、种植成本
    # data2: 所有作物编号对应的预期销售量
    # data3: 所有地块对应的地块类型、面积是否种过豆类
    # data4: data1只有豆类的版本
    # data5: 所有地块上一季种植的作物编号
    with open("jzliang.json", 'r', encoding='utf-8') as f:
        jzliang = json.load(f)
    lianglist = {}
    for adict in jzliang[1:]:
        values = list(adict.values())
        lianglist[values[0]] = values[1]
    with open("jzshu.json", 'r', encoding='utf-8') as f:
        jzshu = json.load(f)
    shulist = {}
    for adict in jzshu[1:]:
        values = list(adict.values())
        shulist[values[0]] = values[1]
    leveldict = {**lianglist, **shulist}
    pealiang = [1, 2, 3, 4, 5]
    peashu = [17, 18, 19]
    bonus = {3: 0.7, 2: 0.3, 1: 0.1}
    # 第一季
    # 降价后利润 (作物编号、地块编号)：降价后利润
    descenddict1 = {}
    # 亩产量 (作物编号、地块编号)：亩产量
    muchan1 = {}
    # 优先队列 全 亩利润 作物编号 地块编号
    pqall1 = []
    # 字典形式的data1_1
    alldict1 = {}
    # 间作编号（100+）
    newid = 100
    newdict1 = {}
    for data in data1_1[1:]:
        values = list(data.values())
        alldict1[(values[0], values[1])] = [values[2], values[3], values[4]]
        descenddict1[(values[0], values[1])] = values[2] / 2 * values[3] - values[4]
        muchan1[(values[0], values[1])] = values[3]
        heapq.heappush(pqall1, (-(values[2] * values[3] - values[4]), values[0], values[1]))

    # 构建间作组合并加入优先队列
    # 粮食
    for liang, level in lianglist.items():
        for pea in pealiang:
            for grid in [1, 2, 3]:
                liangstatus = alldict1[(liang, grid)]
                liangpro = liangstatus[0] * liangstatus[1] * (1 + bonus[level])- liangstatus[2]
                peastatus = alldict1[(pea, grid)]
                peapro = peastatus[0] * peastatus[1] * (1 + 0.1) - peastatus[2]
                pro = (liangpro + peapro) / 2
                newdict1[newid] = [liang, pea]
                heapq.heappush(pqall1, (-pro, newid, grid))
                newid += 1
    # 蔬菜
    for shu, level in shulist.items():
        for pea in peashu:
            for grid in [4, 5, 6]:
                shustatus = alldict1[(shu, grid)]
                shupro = shustatus[0] * shustatus[1] * (1 + bonus[level])- shustatus[2]
                peastatus = alldict1[(pea, grid)]
                peapro = peastatus[0] * peastatus[1] * (1 + 0.1) - peastatus[2]
                pro = (shupro + peapro) / 2
                newdict1[newid] = [shu, pea]
                heapq.heappush(pqall1, (-pro, newid, grid))
                newid += 1

    # 第二季
    # 降价后利润 (作物编号、地块编号)：降价后利润
    descenddict2 = {}
    # 亩产量 (作物编号、地块编号)：亩产量
    muchan2 = {}
    # 优先队列 全 亩利润 作物编号 地块编号
    pqall2 = []
    # 字典形式的data1_2
    alldict2 = {}
    # 间作编号（100+）
    newid = 100
    newdict2 = {}
    for data in data1_2[1:]:
        values = list(data.values())
        alldict2[(values[0], values[1])] = [values[2], values[3], values[4]]
        descenddict2[(values[0], values[1])] = values[2] / 2 * values[3] - values[4]
        muchan2[(values[0], values[1])] = values[3]
        heapq.heappush(pqall2, (-(values[2] * values[3] - values[4]), values[0], values[1]))

    # 间作组合
    # 蔬菜
    for shu, level in shulist.items():
        for pea in peashu:
            for grid in [6]:
                shustatus = alldict2[(shu, grid)]
                shupro = shustatus[0] * shustatus[1] * (1 + bonus[level]) - shustatus[2]
                peastatus = alldict2[(pea, grid)]
                peapro = peastatus[0] * peastatus[1] * (1 + 0.1) - peastatus[2]
                pro = (shupro + peapro) / 2
                newdict2[newid] = [shu, pea]
                heapq.heappush(pqall2, (-pro, newid, grid))
                newid += 1

    # 豆类列表
    pealist = set()
    for data in data4[1:]:
        values = list(data.values())
        pealist.add(values[0])
    # 预期销售量
    maxsold = {}
    for data in data2[1:]:
        values = list(data.values())
        maxsold[values[0]] = values[1]
    maxsold_2 = copy.deepcopy(maxsold)
    # {地块类型1：[（地块1，面积）……]……}
    ground = defaultdict(list)
    for data in data3[1:]:
        values = list(data.values())
        ground[values[1]].append([values[0], values[2]])
    if year == 2023:
        # 是否种豆
        hvpea = {}
        for data in data3[1:]:
            values = list(data.values())
            hvpea[values[0]] = values[3]
    else:
        with open(os.path.join(directory, str(year), "hvpea.json"), 'r', encoding='utf-8') as f:
            hvpea = json.load(f)

    season1pea = defaultdict(int)
    season2pea = defaultdict(int)
    changed = defaultdict(int)
    profit1 = greedysolve(1, muchan1, pqall1, descenddict1, get_row_number1, alldict1, newdict1)
    profit2 = greedysolve(2, muchan2, pqall2, descenddict2, get_row_number2, alldict2, newdict2)
    year += 1
    if str(year) not in os.listdir(directory):
        os.mkdir(os.path.join(directory, str(year)))

    np.random.seed(year)
    percent = {}
    for i in range(1, 42):
        if i in list(range(17, 38)):
            percent[i] = (100 + 5 + np.random.randn()) / 100
        elif i in list(range(39, 41)):
            percent[i] = (100 - 1 - 4 * np.random.random()) / 100
        elif i == 41:
            percent[i] = (100 - 5 - np.random.randn()) / 100
        else:
            percent[i] = 1.

    for adict in data2[1:]:
        if adict['C0'] in [6, 7]:
            adict['C1'] *= (100 + 5 + 5 * np.random.random()) / 100
        else:
            adict['C1'] *= (100 - 5 + 10 * np.random.random()) / 100
        adict['C1'] *= 1. - elasticity[adict["C0"]] * (percent[adict["C0"]] - 1)
    for adict in data1_1[1:]:
        adict['C3'] *= (100 - 10 + 20 * np.random.random()) / 100
        adict['C4'] *= (100 + 5 + np.random.randn()) / 100
        adict['C2'] *= percent[adict['C0']]
    for adict in data1_2[1:]:
        adict['C3'] *= (100 - 10 + 20 * np.random.random()) / 100
        adict['C4'] *= (100 + 5 + np.random.randn()) / 100
        adict['C2'] *= percent[adict['C0']]
    with open(os.path.join(directory, str(year), "1_1.json"), 'w', encoding='utf-8') as f:
        json.dump(data1_1, f, ensure_ascii=False)
    with open(os.path.join(directory, str(year), "1_2.json"), 'w', encoding='utf-8') as f:
        json.dump(data1_2, f, ensure_ascii=False)
    with open(os.path.join(directory, str(year), "2.json"), 'w', encoding='utf-8') as f:
        json.dump(data2, f, ensure_ascii=False)
    with open(os.path.join(directory, str(year), "3.json"), 'w', encoding='utf-8') as f:
        json.dump(data3, f, ensure_ascii=False)
    with open(os.path.join(directory, str(year), "4.json"), 'w', encoding='utf-8') as f:
        json.dump(data4, f, ensure_ascii=False)
    with open(os.path.join(directory, str(year), "hvpea.json"), 'w', encoding='utf-8') as f:
        json.dump(hvpea, f, ensure_ascii=False)
    with open(os.path.join(directory, str(year), "checkhvwhat.json"), 'w', encoding='utf-8') as f:
        json.dump(checkhvwhat, f, ensure_ascii=False)
    with open(os.path.join(directory, str(year), "log.txt"), 'w', encoding='utf-8') as f:
        json.dump("总利润：{}".format(profit1 + profit2), f, ensure_ascii=False)
    print(str(year) + "总利润：{}".format(profit1 + profit2))
    totalprofit += profit1 + profit2
print(totalprofit)
with open('log.txt', 'w', encoding='utf-8') as f:
    f.write(str(totalprofit))
workbook.save('result3.xlsx')
