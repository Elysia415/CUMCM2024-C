import json
import os
import heapq
import copy
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def greedysolve(muchan, pqall, get_row_number, season):
    _ground = copy.deepcopy(ground)
    totalprofit = 0
    checkwhat = defaultdict(list)
    for place, plantlist in checkhvwhat.items():
        for plant in plantlist:
            checkwhat[place].append(plant[0])
    newhvwhat = defaultdict(list)
    while len(pqall) > 0:
        profit, pldi, grid = pqall[0]
        profit = -profit
        # 是否种下
        planted = False
        for i in range(len(_ground[grid])):
            if maxsold[pldi] / muchan[(pldi, grid)] <= limitminarea[grid]:
                planted = True
                # 超过预期销售量直接滞销，移出队列
                heapq.heappop(pqall)
                break
            place, area = _ground[grid][i]
            # 关于豆类的约束
            if hvpea[place] >= 2 and pldi not in pealist:
                continue
            # 关于上一季种植作物的约束
            if pldi in checkwhat[place]:
                continue
            if area <= limitminarea[grid]:
                continue
            planted = True
            # 贪心,尽量种最多
            plantarea = min(area, maxsold[pldi] / muchan[(pldi, grid)])
            totalprofit += plantarea * profit
            _ground[grid][i][1] -= plantarea
            maxsold[pldi] -= plantarea * muchan[(pldi, grid)]
            newhvwhat[place].append([pldi, plantarea])
        if not planted:
            heapq.heappop(pqall)
            continue
    # 更新豆类约束
    if season == 1:
        for place, plantlist in newhvwhat.items():
            p = 0
            for pldi, plantarea in plantlist:
                if pldi in pealist:
                    p = 1
                    break
            season1pea[place] = p
    else:
        for place, plantlist in newhvwhat.items():
            p = 0
            for pldi, plantarea in plantlist:
                if pldi in pealist:
                    p = 1
                    break
            season2pea[place] = p
        for place in hvpea.keys():
            if season1pea[place] or season2pea[place]:
                hvpea[place] = 0
            else:
                hvpea[place] += 1
    for place in newhvwhat:
        finalhvwhat[place] += newhvwhat[place]
        checkhvwhat[place] = newhvwhat[place]
    if season == 1:
        for grid in [1, 2, 3, 4, 5, 6]:
            for place, area in _ground[grid]:
                if place not in newhvwhat:
                    checkhvwhat[place] = []
    # 写入result表格
    sheetname = str(year + 1)
    sheet = workbook[sheetname]
    for place, plantlist in newhvwhat.items():
        for pldi, plantarea in plantlist:
            col = get_column_letter(pldi + 2)
            row = get_row_number[place]
            cell_address = f"{col}{row}"
            sheet[cell_address] = plantarea
    return totalprofit


workbook = load_workbook('result.xlsx')
with open("get_row_number1.json", 'r', encoding='utf-8') as f:
    raw_data = json.load(f)
get_row_number1 = {}
for row_dict in raw_data:
    number, place = row_dict.values()
    get_row_number1[place] = number
with open("get_row_number2.json", 'r', encoding='utf-8') as f:
    raw_data = json.load(f)
get_row_number2 = {}
for row_dict in raw_data:
    number, place = row_dict.values()
    get_row_number2[place] = number
year = 2023
totalprofit = 0
directory = os.getcwd()

limitminarea = {1: 1, 2: 1, 3: 1, 4: 1, 5: 0.1, 6: 0.1}
while year < 2030:
    with open(os.path.join(directory, str(year), "1_1.json"), 'r', encoding='utf-8') as f:
        data1_1 = json.load(f)
    with open(os.path.join(directory, str(year), "1_2.json"), 'r', encoding='utf-8') as f:
        data1_2 = json.load(f)
    with open(os.path.join(directory, str(year), "2.json"), 'r', encoding='utf-8') as f:
        data2 = json.load(f)
    with open(os.path.join(directory, str(year), "3.json"), 'r', encoding='utf-8') as f:
        data3 = json.load(f)
    with open(os.path.join(directory, str(year), "4.json"), 'r', encoding='utf-8') as f:
        data4 = json.load(f)
    if year == 2023:
        with open(os.path.join(directory, str(year), "5.json"), 'r', encoding='utf-8') as f:
            data5 = json.load(f)
    if year != 2023:
        with open(os.path.join(directory, str(year), "checkhvwhat.json"), 'r', encoding='utf-8') as f:
            checkhvwhat = json.load(f)
    else:
        checkhvwhat = defaultdict(list)
        for data in data5:
            values = list(data.values())
            checkhvwhat[values[0]].append([values[1], 0])

    # data1_1/1_2: 每一季（作物编号，地块类型编号）所对应的单价、亩产量、种植成本
    # data2: 所有作物编号对应的预期销售量
    # data3: 所有地块对应的地块类型、面积是否种过豆类
    # data4: data1只有豆类的版本
    # data5: 所有地块上一季种植的作物编号

    # 第一季
    # 亩产量 (作物编号、地块编号)：亩产量
    muchan1 = {}
    # 优先队列 全 亩利润 作物编号 地块编号
    pqall1 = []
    for data in data1_1[1:]:
        values = list(data.values())
        muchan1[(values[0], values[1])] = values[3]
        heapq.heappush(pqall1, (-(values[2] * values[3] - values[4]), values[0], values[1]))

    # 第二季
    # 亩产量 (作物编号、地块编号)：亩产量
    muchan2 = {}
    # 优先队列 全 亩利润 作物编号 地块编号
    pqall2 = []
    for data in data1_2[1:]:
        values = list(data.values())
        muchan2[(values[0], values[1])] = values[3]
        heapq.heappush(pqall2, (-(values[2] * values[3] - values[4]), values[0], values[1]))

    # 豆类列表
    pealist = set()
    for data in data4[1:]:
        values = list(data.values())
        pealist.add(values[0])
    # 预期销售量
    maxsold = {}
    for data in data2[1:]:
        values = list(data.values())
        maxsold[values[0]] = values[1]
    # {地块类型1：[（地块1，面积）……]……}
    ground = defaultdict(list)
    for data in data3[1:]:
        values = list(data.values())
        ground[values[1]].append([values[0], values[2]])
    if year == 2023:
        # 是否种豆
        hvpea = {}
        for data in data3[1:]:
            values = list(data.values())
            hvpea[values[0]] = values[3]
    else:
        with open(os.path.join(directory, str(year), "hvpea.json"), 'r', encoding='utf-8') as f:
            hvpea = json.load(f)

    finalhvwhat = defaultdict(list)
    season1pea = defaultdict(int)
    season2pea = defaultdict(int)
    changed = defaultdict(int)
    profit1 = greedysolve(muchan1, pqall1, get_row_number1, 1)
    profit2 = greedysolve(muchan2, pqall2, get_row_number2, 2)
    year += 1
    if str(year) not in os.listdir(directory):
        os.mkdir(os.path.join(directory, str(year)))
    with open(os.path.join(directory, str(year), "1_1.json"), 'w', encoding='utf-8') as f:
        json.dump(data1_1, f, ensure_ascii=False)
    with open(os.path.join(directory, str(year), "1_2.json"), 'w', encoding='utf-8') as f:
        json.dump(data1_2, f, ensure_ascii=False)
    with open(os.path.join(directory, str(year), "2.json"), 'w', encoding='utf-8') as f:
        json.dump(data2, f, ensure_ascii=False)
    with open(os.path.join(directory, str(year), "3.json"), 'w', encoding='utf-8') as f:
        json.dump(data3, f, ensure_ascii=False)
    with open(os.path.join(directory, str(year), "4.json"), 'w', encoding='utf-8') as f:
        json.dump(data4, f, ensure_ascii=False)
    with open(os.path.join(directory, str(year), "hvpea.json"), 'w', encoding='utf-8') as f:
        json.dump(hvpea, f, ensure_ascii=False)
    with open(os.path.join(directory, str(year), "checkhvwhat.json"), 'w', encoding='utf-8') as f:
        json.dump(checkhvwhat, f, ensure_ascii=False)
    with open(os.path.join(directory, str(year), "log.txt"), 'w', encoding='utf-8') as f:
        json.dump("总利润：{}".format(profit1 + profit2), f, ensure_ascii=False)
    print(str(year) + "总利润：{}".format(profit1 + profit2))
    totalprofit += profit1 + profit2
print(totalprofit)
with open('log.txt', 'w', encoding='utf-8') as f:
    f.write(str(totalprofit))
workbook.save('result1_1.xlsx')
